from airflow.providers.postgres.hooks.postgres import PostgresHook
from openpyxl import load_workbook


def insert_rows_to_postgres(pg_hook, insert_sql, data_rows):
    with pg_hook.get_conn() as conn:
        with conn.cursor() as cursor:
            for row in data_rows:
                cursor.execute(insert_sql, row)
        conn.commit()
        

def read_excel_data(filepath):
    wb = load_workbook(filename=filepath)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    return rows[1:]


def read_xlsx_and_load_to_postgres(**kwargs):
    try: 
        data_rows = read_excel_data('/opt/airflow/dags/data/users_data.xlsx')

        pg_hook = PostgresHook(postgres_conn_id='my_postgres_conn')

        insert_sql = """
            INSERT INTO raw.users (name, surname, city, age, card_number)
            VALUES (%s, %s, %s, %s, %s);
        """
        insert_rows_to_postgres(pg_hook, insert_sql, data_rows)
        
        print("[Данные вставились]")

    except Exception as e:
        print(f"Ошибка загрузки данных: {e}")


def read_orders_and_load_to_postgres(**kwargs):
    try:
        data_rows = read_excel_data('/opt/airflow/dags/data/orders_data.xlsx')

        pg_hook = PostgresHook(postgres_conn_id='my_postgres_conn')

        insert_sql = """
            INSERT INTO raw.orders (id, name, quantity, price, total_price, card_number)
            VALUES (%s, %s, %s, %s, %s, %s);
        """

        insert_rows_to_postgres(pg_hook, insert_sql, data_rows)
        
        print("[Данные вставились]")

    except Exception as e:
        print(f"Ошибка загрузки данных: {e}")

def read_deliveries_and_load_to_postgres(**kwargs):
    try:
        data_rows = read_excel_data('/opt/airflow/dags/data/deliveries_data.xlsx')

        pg_hook = PostgresHook(postgres_conn_id='my_postgres_conn')

        insert_sql = """
            INSERT INTO raw.deliveries (delivery_number, order_id, goods_array, delivery_company, delivery_price, courier_name, courier_phone, start_date, end_date, city, warehouse, warehouse_address)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
        """
        insert_rows_to_postgres(pg_hook, insert_sql, data_rows)
        
        print("[Данные вставились]")

    except Exception as e:
        print(f"Ошибка загрузки данных: {e}")

def get_users_total_orders_price():
    try:
        pg_hook = PostgresHook(postgres_conn_id='my_postgres_conn')
        with pg_hook.get_conn() as conn:
            with conn.cursor() as cursor:

                select_sql = """
                        select 
                            u.name,
                            u.surname,
                            SUM(o.total_price),
                            o.card_number
                        from raw.users u 
                        join raw.orders o on u.card_number = o.card_number
                        group by u.name, u.surname, o.card_number;
                    """
                cursor.execute(select_sql)
                result = cursor.fetchall()

                insert_sql = """
                        insert into data_mart.users_total_orders (name, surname, total_price, card_number)
                        values (%s, %s, %s, %s)
                    """

                for row in result:
                    cursor.execute(insert_sql, row)

            conn.commit()

        print("[Данные вставились]")

    except Exception as e:
        print(f"Ошибка при загрузке данных: {e}")